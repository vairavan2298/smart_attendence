# import openpyxl module 
from datetime import datetime
import openpyxl 

def dbval(val):

# Call a Workbook() function of openpyxl 
# to create a new blank Workbook object 
					wb = openpyxl.Workbook() 

					# Get workbook active sheet 
					# from the active attribute 
					sheet = wb.active 

					# Cell objects also have row, column 
					# and coordinate attributes that provide 
					# location information for the cell. 

					# Note: The first row or column integer 
					# is 1, not 0. Cell object is created by 
					# using sheet object's cell() method. 

					now = datetime.now()
			 
				# print("now =", now)
					dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
					c1 = sheet.cell(row =1 ,column = 1) 

					# writing values to cells 
					c1.value = val

					c2 = sheet.cell(row= 1, column = 2) 
					c2.value = dt_string

					# Once have a Worksheet object, one can 
					# access a cell object by its name also. 
					# A2 means column = 1 & row = 2. 
					# c3 = sheet.cell(row = 2,column = 1)
				    
				 #    c3.value= val

					# # # B2 means column = 2 & row = 2. 
					# c4 = sheet['B2'] 
					# c4.value = dt_string

					# Anytime you modify the Workbook object 
					# or its sheets and cells, the spreadsheet 
					# file will not be saved until you call 
					# the save() workbook method. 
					wb.save("demo.xlsx") 
